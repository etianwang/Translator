import sys
import os
import sys
import os
import time
import json
from datetime import datetime
from PyQt5 import QtWidgets, QtGui, QtCore
from PyQt5.QtWidgets import QApplication, QFileDialog, QMessageBox, QDialog, QVBoxLayout, QLineEdit, QPushButton
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.cell.cell import MergedCell
from copy import copy
from .translator import translate_cell_text
from .translator import load_term_dict
opened_windows = {}
def resource_path(relative_path):
    """兼容 PyInstaller 打包和正常运行的资源路径"""
    try:
        base_path = sys._MEIPASS  # 打包后执行环境
    except AttributeError:
        base_path = os.path.dirname(os.path.abspath(__file__))  # 脚本真实位置

    return os.path.join(base_path, relative_path)

CONFIG_PATH = ".translator_config.json"

def load_api_keys():
    if os.path.exists(CONFIG_PATH):
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"deepl": ""}

def save_api_keys(keys):
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(keys, f, indent=2)

class ApiKeyDialog(QDialog):
    def __init__(self, current_key, parent=None):
        super().__init__(parent)
        self.setWindowTitle("输入 DeepL API Key")
        self.resize(400, 120)
        layout = QVBoxLayout(self)
        self.input = QLineEdit(self)
        self.input.setText(current_key)
        self.input.setPlaceholderText("请输入 DeepL 的 API Key")
        layout.addWidget(self.input)
        self.ok_btn = QPushButton("确定", self)
        self.ok_btn.clicked.connect(self.accept)
        layout.addWidget(self.ok_btn)

    def get_key(self):
        return self.input.text().strip()

class ExcelTranslatorUI(QtWidgets.QWidget):
    def mousePressEvent(self, event):
        if event.button() == QtCore.Qt.MouseButton.LeftButton:
            self._mouse_pos = event.globalPos() - self.frameGeometry().topLeft()
            event.accept()

    def mouseMoveEvent(self, event):
        if event.buttons() == QtCore.Qt.MouseButton.LeftButton:
            self.move(event.globalPos() - self._mouse_pos)
            event.accept()

    def __init__(self):
        super().__init__()
        self.setFixedSize(800, 860)
        self.setWindowFlags(QtCore.Qt.WindowType.FramelessWindowHint)
        self.setAttribute(QtCore.Qt.WidgetAttribute.WA_TranslucentBackground)
        self.api_keys = load_api_keys()
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Honsen Excel 翻译器")

        background = QtWidgets.QLabel(self)
        background.setStyleSheet("background-color: rgba(255, 255, 255, 150); border-radius: 20px;")
        background.setGeometry(0, 0, 800, 860)
        background.lower()

        container = QtWidgets.QWidget(self)
        container.setObjectName("container")
        container.setGeometry(20, 20, 760, 820)
        container.setStyleSheet("""
            #container {
                background-color: rgba(255, 255, 255, 100);
                border-radius: 20px;
                padding: 20px;
            }
            QLineEdit, QComboBox, QTextEdit {
                padding: 6px;
                border: 1px solid #ccc;
                border-radius: 10px;
                background-color: rgba(255, 255, 255, 200);
            }
            QPushButton {
                background-color: rgba(70, 130, 180, 0.8);
                color: white;
                padding: 8px;
                border: none;
                border-radius: 10px;
            }
            QPushButton:hover {
                background-color: rgba(30, 144, 255, 0.8);
            }
        """)

        layout = QtWidgets.QVBoxLayout(container)

        close_btn = QtWidgets.QPushButton("✕")
        close_btn.setFixedSize(32, 32)
        close_btn.setStyleSheet("background-color: transparent; color: #333; font-size: 18px;")
        close_btn.clicked.connect(self.close)
        layout.addWidget(close_btn, alignment=QtCore.Qt.AlignmentFlag.AlignRight)

        title_label = QtWidgets.QLabel("Honsen Excel 翻译器（DEEPL版）")
        title_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        title_label.setStyleSheet("font-size: 26px; font-weight: bold;")
        layout.addWidget(title_label)

        file_layout = QtWidgets.QHBoxLayout()
        self.file_input = QtWidgets.QLineEdit()
        browse_btn = QtWidgets.QPushButton("浏览 Excel")
        browse_btn.clicked.connect(self.select_file)
        file_layout.addWidget(self.file_input)
        file_layout.addWidget(browse_btn)

        self.lang_combo = QtWidgets.QComboBox()
        self.lang_combo.addItems(["zh-fr", "fr-zh", "en-fr", "fr-en"])

        key_btn = QtWidgets.QPushButton("设置 DeepL API Key")
        key_btn.clicked.connect(self.set_api_key)

        self.translate_btn = QtWidgets.QPushButton("开始翻译")
        self.translate_btn.clicked.connect(self.translate_excel)

        self.log_output = QtWidgets.QTextEdit()
        self.log_output.setReadOnly(True)

        layout.addLayout(file_layout)
        layout.addWidget(QtWidgets.QLabel("语言方向:"))
        layout.addWidget(self.lang_combo)
        layout.addWidget(key_btn)
        layout.addWidget(self.translate_btn)
        layout.addWidget(QtWidgets.QLabel("日志输出:"))
        layout.addWidget(self.log_output)

    def set_api_key(self):
        current = self.api_keys.get("deepl", "")
        dialog = ApiKeyDialog(current, self)
        if dialog.exec() == QDialog.Accepted:
            self.api_keys["deepl"] = dialog.get_key()
            save_api_keys(self.api_keys)

    def log(self, msg):
        self.log_output.append(msg)
        QtWidgets.QApplication.processEvents()

    def select_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "选择 Excel 文件", "", "Excel 文件 (*.xlsx)")
        if file_path:
            self.file_input.setText(file_path)

    def translate_excel(self):
        file_path = self.file_input.text().strip()
        if not file_path:
            self.log("❗ 请选择 Excel 文件。")
            return

        lang_pair = self.lang_combo.currentText()
        save_api_keys(self.api_keys)

        self.translate_btn.setEnabled(False)
        self.translate_btn.setText("处理中...")

        self.thread = ExcelTranslateThread(file_path, lang_pair, self.api_keys)
        self.thread.log_signal.connect(self.log)
        self.thread.done_signal.connect(self.on_finished)
        self.thread.start()

    def on_finished(self, output_path):
        QMessageBox.information(self, "翻译完成", f"文件已保存至：{output_path}")
        os.startfile(output_path)
        self.translate_btn.setEnabled(True)
        self.translate_btn.setText("开始翻译")

class ExcelTranslateThread(QtCore.QThread):
    log_signal = QtCore.pyqtSignal(str)
    done_signal = QtCore.pyqtSignal(str)

    def __init__(self, file_path, lang_pair, apikeys):
        super().__init__()
        self.file_path = file_path
        self.lang_pair = lang_pair
        self.apikeys = apikeys

    def run(self):
        _, term_warning = load_term_dict()
        if term_warning:
            self.log_signal.emit(term_warning)
        try:
            self.log_signal.emit("📄 正在处理 Excel 文件...")
            wb = load_workbook(self.file_path)

            for sheet in wb.worksheets:
                self.log_signal.emit(f"➡️ 正在处理工作表：{sheet.title}...")
                max_col = sheet.max_column
                max_row = sheet.max_row
                merged_ranges = list(sheet.merged_cells.ranges)

                for col in reversed(range(1, max_col + 1)):
                    insert_at = col + 1
                    sheet.insert_cols(insert_at)
                    col_letter = get_column_letter(col)
                    new_col_letter = get_column_letter(insert_at)
                    sheet.column_dimensions[new_col_letter].width = sheet.column_dimensions[col_letter].width

                    # 复制合并单元格结构（仅垂直合并的一列）
                    for merge in merged_ranges:
                        if merge.min_col == col and merge.max_col == col:
                            sheet.merge_cells(
                                start_row=merge.min_row,
                                end_row=merge.max_row,
                                start_column=insert_at,
                                end_column=insert_at
                            )

                    for row in range(1, max_row + 1):
                        if row % 10 == 0 or row == max_row:
                            self.log_signal.emit(f"    正在处理第 {row}/{max_row} 行...")
                        source = sheet.cell(row=row, column=col)
                        target = sheet.cell(row=row, column=insert_at)

                        # 跳过合并单元格子项，避免写入 MergedCell 报错
                        if any(
                            merge.min_row <= row <= merge.max_row and
                            merge.min_col == col and merge.max_col == col and
                            row != merge.min_row
                            for merge in merged_ranges
                        ):
                            continue

                        target.value = source.value
                        target.font = copy(source.font)
                        target.border = copy(source.border)
                        target.fill = copy(source.fill)
                        target.alignment = copy(source.alignment)
                        target.number_format = source.number_format

                        if row > 1:
                            # 如果是函数公式（以 = 开头），不翻译，只复制
                            if isinstance(source.value, str) and source.value.strip().startswith("="):
                                continue  # ✅ 保留公式原样，已在上方复制过
                            translated = translate_cell_text(
                                text=source.value,
                                engine="deepl",
                                lang_pair=self.lang_pair,
                                apikeys=self.apikeys,
                                log_func=self.log_signal.emit  # ✅ 传入日志方法
                            )
                            target.value = translated
            
            src, dest = self.lang_pair.split('-')
            base_name = os.path.splitext(os.path.basename(self.file_path))[0]
            time_str = datetime.now().strftime("%Hh%M_%d%m%y")
            out_name = f"{dest}_{base_name}_{time_str}.xlsx"
            out_path = os.path.join(os.path.dirname(self.file_path), out_name)
            wb.save(out_path)
            self.log_signal.emit("✅ 所有工作表已翻译完成。")
            self.done_signal.emit(out_path)
            

        except Exception as e:
            self.log_signal.emit(f"❌ 错误：{e}")

def run_excel_gui():
    app = QtWidgets.QApplication.instance()
    if app is None:
        app = QtWidgets.QApplication(sys.argv)
    window = ExcelTranslatorUI()
    window.show()
    opened_windows["excel"] = window

if __name__ == "__main__":
    run_excel_gui()

